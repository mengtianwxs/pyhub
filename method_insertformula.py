# __*__encoding:utf8__*__

import openpyxl
import openpyxl.worksheet


wb = openpyxl.load_workbook("files/aabb.xlsx")
st = wb.active

row_max = st.max_row
print(row_max)
cell0=[]
cell1 = []
cell2 = []
cell3 = []
cell4=[]

for i in range(1, row_max):
    btxt0=st.cell(row=i,column=1).value
    btxt1 = st.cell(row=i, column=2).value
    btxt2 = st.cell(row=i, column=2).value
    btxt3 = st.cell(row=i, column=2).value
    btxt4=st.cell(row=i,column=2).value

    if(btxt0==1):
        cell0.append(i)
    if (btxt1 == u"小计"):
        cell1.append(i)
    if (btxt2 == u"辅材"):
        cell2.append(i)
    if (btxt3 == u"单台合计"):
        cell3.append(i)
    if (btxt4 ==u"总计"):
        cell4.append(i)


for i in range(0, len(cell1) - 1):
    #小计
    st['G'+str(cell1[i])].value="=SUM(G"+str(cell0[i])+":G"+str(cell1[i]-1)+")"

    st['G' + str(cell3[i])].value = "=SUM(G" + str(cell1[i]) + ":G" + str(cell2[i]) + ")"
    st['G' + str(cell4[i])].value = "=G" + str(cell3[i]) + "*E" + str(cell4[i])


wb.save("files/new_aabb.xlsx")
